from tkinter import *
from tkinter.filedialog import askopenfilename
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime

import win32com.client as win32

FORNECEDORES = {

    'ABB': {
        'UAT': {
            'disjuntor': True,
            'chaves': True,
            'tis': True,
            'para-raios': True,
            'bando de capacitor': True,
            'filtro de harmônicos': False,
            'bobina de bloqueio': False,
            'transformador': True
        },
        'AT': {
            'disjuntor': True,
            'chaves': True,
            'tis': True,
            'para-raios': True,
            'banco de capacitor': True,
            'filtro de harmônicos': False,
            'bobina de bloqueio': False,
            'transformador': True
        },
        'MT': {
            'disjuntor': True,
            'chaves': False,
            'tis': False,
            'para-raios': False,
            'banco de capacitor': True,
            'filtro de harmônicos': True,
            'bobina de bloqueio': False,
            'cububículo de média tensão': True,
            'transformado de serviço auxiliar': False,
            'resistor de aterramento': False,
            'transformador': True
        },
        'BT': {
            'banco de baterias': False,
            'retificador': False,
            'gerador': False
        },
        'contatos':
        [
            {
                'nome': 'Márcio',
                'sobrenome': 'Tanese',
                'email': 'andreversiani@visionsistemas.com.br',
                'tratamento': 'para-raiosezado',
            },
            {
                'nome': 'Gabriel',
                'sobrenome': 'Queiroz',
                'email': 'andreversiani01@gmail.com',
                'tratamento': 'para-raiosezado'
            }
        ]
    },
    'WEG': {
        'UAT': {
            'disjuntor': False,
            'chaves': False,
            'tis': False,
            'para-raios': False,
            'bando de capacitor': False,
            'filtro de harmônicos': False,
            'bobina de bloqueio': False,
            'transformador': False
        },
        'AT': {
            'disjuntor': False,
            'chaves': False,
            'tis': False,
            'para-raios': False,
            'banco de capacitor': False,
            'filtro de harmônicos': False,
            'bobina de bloqueio': False,
            'transformador': False
        },
        'MT': {
            'disjuntor': False,
            'chaves': False,
            'tis': False,
            'para-raios': False,
            'banco de capacitor': True,
            'filtro de harmônicos': True,
            'bobina de bloqueio': False,
            'cubículo de média tensão': True,
            'transformado de serviço auxiliar': False,
            'resistor de aterramento': False,
            'transformador': False
        },
        'BT': {
            'banco de baterias': True,
            'retificador': True,
            'gerador': True
        },
        'contatos':
        [
            {
                'nome': 'Márcio',
                'sobrenome': 'Tanese',
                'email': 'lumini2@hotmail.com.br',
                'tratamento': 'prezado',
            }
        ]
    }
}


DESCRICAO_COLUMN = 'B'
QTD_COLUMN = 'AP'
VOLTAGE_COLUMN = 'E'

SHEETS = ['transformador', 'disjuntor', 'chaves', 'tis', 'para-raios', 'banco de capacitor',
          'resistor de aterramento', 'banco de baterias', 'retificador', 'filtro de harmônicos',
          'transformador de serviço aux', 'cubículo de média tensão', 'bobina de bloqueio']


class Equipamento:

    def __init__(self, tensao, type, qtd, descricao, et=None):
        self.et = et
        self.tensao = tensao
        self.type = type
        self.qtd = qtd
        self.descricao = descricao


class Emailer:
    def __init__(self, receiver, subject, body, attachments=None):
        self.receiver = receiver
        self.subject = subject
        self.body = body
        self.attachments = attachments

    def prepare_emails(self):
        outlook = win32.Dispatch('outlook.application')
        email = outlook.CreateItem(0)
        email.Display()
        bodystart = re.search("<body.*?>", email.HTMLBody)
        email.HTMLBody = re.sub(bodystart.group(), bodystart.group() +
                                self.body, email.HTMLBody)
        email.To = self.receiver
        email.Subject = self.subject
        if self.attachments:
            for attachments in self.attachments:
                email.Attachments.Add(attachments)
        email.CC = 'comercialenergia@visionsistemas.com.br'


class Fornecedor:
    def __init__(self, information):
        self.information = information


def get_et():

    window = Tk()
    window.withdraw()
    filename = askopenfilename()
    et = filename
    window.destroy()
    if et != '':
        return et
    return None


def find_voltage_class(num):
    if num in range(0, 1000):
        return 'BT'
    if num in range(1000, 40000):
        return 'MT'
    if num in range(40000, 250000):
        return 'AT'
    if num in range(250000, 600000):
        return 'UAT'


def get_data_from_equipamentos_sheet():
    window = Tk()
    window.withdraw()
    filename = askopenfilename()
    wb = load_workbook(filename, data_only=True)
    window.destroy()
    equipamentos = []
    need_attachments = False
    for ws_name in wb.sheetnames:
        if ws_name.lower() in SHEETS:
            current_working_sheet = wb[ws_name]
            for row in range(3, current_working_sheet.max_row):
                current_qtd_cell_value = current_working_sheet[f'{QTD_COLUMN}{row}'].value
                if current_qtd_cell_value:
                    need_attachments = True

            if need_attachments:
                et = get_et()
                print(f'Insira a ET: {ws_name}')

            for row in range(3, current_working_sheet.max_row):
                current_qtd_cell_value = current_working_sheet[f'{QTD_COLUMN}{row}'].value
                current_descricao_cell_value = current_working_sheet[f'{DESCRICAO_COLUMN}{row}'].value
                current_voltage_cell_value = current_working_sheet[f'{VOLTAGE_COLUMN}{row}'].value
                if current_qtd_cell_value:
                    need_attachments = True
                    group = ws_name.lower()
                    descricao = current_descricao_cell_value
                    voltage_class = find_voltage_class(
                        current_voltage_cell_value * 1000)
                    qtd = current_qtd_cell_value
                    new_equipamento = Equipamento(
                        voltage_class, group, qtd, descricao, et)
                    equipamentos.append(new_equipamento)

        need_attachments = False

    return equipamentos


equipamentos = get_data_from_equipamentos_sheet()

resumo = {}
for fornecedor in FORNECEDORES:
    for equipamento in equipamentos:
        try:
            validator = FORNECEDORES[fornecedor][equipamento.tensao][equipamento.type]
        except Exception as e:
            pass
        if validator:
            if fornecedor not in resumo:
                resumo.update(
                    {fornecedor: []})
            resumo[fornecedor].append(equipamento)

for fornecedor in resumo:
    contatos = FORNECEDORES[fornecedor]['contatos']
    hora = datetime.datetime.now().ctime()
    hora = int(hora[11:13])

    if hora in range(0, 13):
        introducao = 'bom dia'
    if hora in range(13, 18):
        introducao = 'boa Tarde'
    if hora in range(18, 24):
        introducao = 'boa noite'

    if len(contatos) > 1:
        tratamento = 'Prezados'
    if len(contatos) == 1:
        tratamento = contatos[0]['tratamento'].capitalize()
        introducao = contatos[0]['nome'] + introducao

    text = ''
    item = 1
    ets = []
    for eq in resumo[fornecedor]:
        if eq.et:
            ets.append(eq.et)

        TEXT_COLOR = '#3c4064'
        TABLE_STYLE = 'border-collapse:collapse; margin: 5px; font-size: 14px; width: 500px;'
        TR_STYLE = 'background-color:#154c79; color: white; font-weight:bold;border: 1px solid; border: 1px solid'
        TH_STYLE = 'padding: 2px 5px; border: 1px solid'
        TD_STYLE = 'text-align:center; padding: 2px 5px; border: 1px solid; background-color: white'

        text += f"""
        <tr'>
            <td style='{TD_STYLE}'>{item}</td>
            <td style='{TD_STYLE}'>{eq.descricao}</td>
            <td style='{TD_STYLE}'>{eq.qtd}</td>
        </tr>
        """
        item += 1
        body = f"""
                <div>
                <p style='color: {TEXT_COLOR}'p>{tratamento}, {introducao}, </p>
                <p style='color: {TEXT_COLOR}'p>Devido à necessidade do setor comercial da Vision Engenharia, solicitamos o orçamento dos equipamentos destacados na tabela a seguir. O anexo contém mais informações a serem consideradas.</p>
                <table style='{TABLE_STYLE}'>
                        <tr style='{TR_STYLE}'>
                            <th style='{TH_STYLE}'>ITEM</th>
                            <th style='{TH_STYLE}'>DESCRIÇÃO</th>
                            <th style='{TH_STYLE}'>QTD</th>
                        </tr>
                    {text}   
                </table>
                    <p style='color: {TEXT_COLOR}'>O prazo máximo para a cotação é segunda-feira, dia 10/01/2022</p>
                    <p style='color: {TEXT_COLOR}'>Observações:</p>
                    <p style='color: {TEXT_COLOR}'>- Caso não seja possível realizar o orçamento dentro do prazo, favor informar em resposta a este e-mail;</p>
                    <p style='color: {TEXT_COLOR}'>- Favor responder a todos os que estão em cópia.</p>
                    <p style='color: {TEXT_COLOR}'>Desde já, a Vision agradece o seu apoio e a sua atenção e nos colocamos à disposição para sanar quaisquer dúvidas sobre o orçamento.</p>
                </div>
            """

    to = []
    for contato in contatos:
        to.append(contato['email'])
    to = ';'.join(to)

    emailer = Emailer(to, 'Cotação', body, ets)
    emailer.prepare_emails()
